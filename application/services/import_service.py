from __future__ import annotations

"""
Importación de ensayos desde archivo, para identificar sin PLC conectado.

Formatos soportados (la detección es automática):

* **Trace de CODESYS / ctrlX** (`.trace.csv`): el export nativo del trace.
  No es un CSV tabular: trae metadata `clave; valor`, y por cada variable un
  bloque `N.Variable; nombre` seguido de filas `; t_ms; valor`.
* **CSV tabular**: primera fila con encabezados, una columna de tiempo y una
  columna por señal. Separador `,` o `;` (detectado).
* **Excel** (`.xlsx`): primera hoja, misma estructura que el CSV tabular.

El servicio solo EXTRAE: devuelve las variables encontradas con sus series y
un resumen para que la vista muestre el modal de mapeo. Decidir qué variable
es actuador/sensor/setpoint es del usuario, igual que con el PLC en vivo.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class ImportedVariable:
    name: str
    values: list[float] = field(default_factory=list)

    def summary(self) -> dict:
        vs = self.values
        n = len(vs)
        preview = vs[:8]

        return {
            "name": self.name,
            "samples": n,
            "min": min(vs) if vs else None,
            "max": max(vs) if vs else None,
            "first": vs[0] if vs else None,
            "last": vs[-1] if vs else None,
            "preview": preview,
            # Una señal que no se mueve no sirve como actuador ni como sensor.
            "constant": n > 1 and abs(max(vs) - min(vs)) < 1e-9,
        }


@dataclass
class ImportedDataset:
    source_name: str
    format: str  # "codesys-trace" | "csv" | "xlsx"
    time_s: list[float] = field(default_factory=list)
    variables: list[ImportedVariable] = field(default_factory=list)

    @property
    def sample_period_s(self) -> float:
        """Mediana de los deltas: robusta a algún hueco suelto en el trace."""
        t = self.time_s
        if len(t) < 2:
            return 0.0
        deltas = sorted(t[i + 1] - t[i] for i in range(len(t) - 1))
        return deltas[len(deltas) // 2]

    def describe(self) -> dict:
        return {
            "source_name": self.source_name,
            "format": self.format,
            "samples": len(self.time_s),
            "duration_s": (self.time_s[-1] - self.time_s[0]) if self.time_s else 0.0,
            "sample_period_s": round(self.sample_period_s, 6),
            "variables": [v.summary() for v in self.variables],
        }

    def variable(self, name: str) -> Optional[ImportedVariable]:
        wanted = _normalize(name)
        for v in self.variables:
            if _normalize(v.name) == wanted:
                return v
        return None


def _normalize(name: str) -> str:
    return (name or "").strip().lower()


# --------------------------------------------------------------------------- #
# Detección y despacho
# --------------------------------------------------------------------------- #


def parse_file(filename: str, content: bytes) -> ImportedDataset:
    """Punto de entrada: detecta el formato y parsea."""
    lower = (filename or "").lower()

    if lower.endswith(".xlsx") or content[:4] == b"PK\x03\x04":
        return _parse_xlsx(filename, content)

    text = content.decode("utf-8-sig", errors="replace")

    # El trace de CODESYS se reconoce por su cabecera y sus bloques.
    if re.search(r"^\d+\.Variable;", text, re.M) or text.startswith("[key]; [value]"):
        return _parse_codesys_trace(filename, text)

    return _parse_tabular_csv(filename, text)


# --------------------------------------------------------------------------- #
# Trace de CODESYS / ctrlX
# --------------------------------------------------------------------------- #


def _parse_codesys_trace(filename: str, text: str) -> ImportedDataset:
    """
    Estructura del export:

        [key]; [value]
        ...metadata...
        0.Variable; DB_HMI.HMI_SP_Local_Automatico
        ...atributos del canal 0...
        0.Data;
        ; 2; 25          <- ; tiempo_ms; valor
        ; 22; 25
        1.Variable; Main_Control.Velocidad_Scaled
        ...

    Los timestamps vienen en milisegundos y cada canal trae los suyos. Se
    exige que coincidan entre canales: si el trace se configuró con distintas
    tasas por canal, mezclarlos sin re-muestrear daría series desalineadas.
    """
    channels: list[tuple[str, list[float], list[float]]] = []  # (nombre, t_ms, valores)
    current_t: Optional[list[float]] = None
    current_v: Optional[list[float]] = None

    var_re = re.compile(r"^\d+\.Variable;\s*(.+?)\s*$")
    row_re = re.compile(r"^;\s*([\d.eE+-]+);\s*([\d.eE+-]+)\s*$")

    for line in text.splitlines():
        m = var_re.match(line)
        if m:
            current_t, current_v = [], []
            channels.append((m.group(1), current_t, current_v))
            continue

        m = row_re.match(line)
        if m and current_t is not None:
            try:
                current_t.append(float(m.group(1)))
                current_v.append(float(m.group(2)))
            except ValueError:
                continue

    channels = [(n, t, v) for n, t, v in channels if t]

    if not channels:
        raise ValueError(
            "El archivo tiene formato de trace de CODESYS pero no se encontró "
            "ningún canal con datos."
        )

    base_t = channels[0][1]

    for name, t, _ in channels[1:]:
        if len(t) != len(base_t) or t[:5] != base_t[:5] or t[-5:] != base_t[-5:]:
            raise ValueError(
                f"Los timestamps del canal '{name}' no coinciden con los de "
                f"'{channels[0][0]}' ({len(t)} vs {len(base_t)} muestras). "
                "Exporta el trace con la misma tasa de muestreo en todos los canales."
            )

    return ImportedDataset(
        source_name=filename,
        format="codesys-trace",
        time_s=[t / 1000.0 for t in base_t],  # ms -> s
        variables=[ImportedVariable(name=n, values=v) for n, _, v in channels],
    )


# --------------------------------------------------------------------------- #
# CSV tabular
# --------------------------------------------------------------------------- #

# Nombres de columna que se reconocen como el eje de tiempo.
TIME_HINTS = ("time", "tiempo", "t_s", "t(s)", "seg", "sec", "ms", "timestamp")


def _looks_like_time(header: str) -> bool:
    h = _normalize(header)
    return any(hint in h for hint in TIME_HINTS)


def _parse_rows(filename: str, fmt: str, headers: list[str], rows: list[list]) -> ImportedDataset:
    """Común a CSV y xlsx: de filas crudas a dataset."""
    if not headers or len(headers) < 2:
        raise ValueError(
            "Se necesitan al menos dos columnas: una de tiempo y una señal."
        )

    # Columna de tiempo: por nombre; si ninguna suena a tiempo, la primera.
    time_idx = next((i for i, h in enumerate(headers) if _looks_like_time(h)), 0)

    columns: dict[int, list[float]] = {i: [] for i in range(len(headers))}

    for row in rows:
        # Fila válida = el tiempo es numérico. Las demás celdas no numéricas
        # se descartan fila completa para mantener las series alineadas.
        try:
            parsed = [float(str(c).replace(",", ".")) for c in row[: len(headers)]]
        except (TypeError, ValueError):
            continue
        if len(parsed) < len(headers):
            continue
        for i, value in enumerate(parsed):
            columns[i].append(value)

    time_values = columns[time_idx]

    if len(time_values) < 2:
        raise ValueError(
            "No se encontraron filas numéricas suficientes. Revisa que el "
            "archivo tenga encabezados en la primera fila y datos numéricos."
        )

    # ¿El tiempo está en ms? Si el paso típico es >= 1 y el header sugiere ms,
    # o el paso es claramente grande para ser segundos, se convierte.
    deltas = sorted(
        time_values[i + 1] - time_values[i] for i in range(len(time_values) - 1)
    )
    dt = deltas[len(deltas) // 2]
    header_is_ms = "ms" in _normalize(headers[time_idx])

    if header_is_ms or dt >= 5.0:
        time_values = [t / 1000.0 for t in time_values]

    variables = [
        ImportedVariable(name=headers[i], values=columns[i])
        for i in range(len(headers))
        if i != time_idx
    ]

    return ImportedDataset(
        source_name=filename,
        format=fmt,
        time_s=time_values,
        variables=variables,
    )


def _parse_tabular_csv(filename: str, text: str) -> ImportedDataset:
    # Separador: el que más aparezca en la primera línea con contenido.
    first = next((ln for ln in text.splitlines() if ln.strip()), "")
    delimiter = ";" if first.count(";") > first.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = [r for r in reader if any(str(c).strip() for c in r)]

    if not all_rows:
        raise ValueError("El archivo está vacío.")

    headers = [str(h).strip() for h in all_rows[0]]
    return _parse_rows(filename, "csv", headers, all_rows[1:])


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #


def _parse_xlsx(filename: str, content: bytes) -> ImportedDataset:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError(
            "Falta la librería openpyxl para leer .xlsx (pip install openpyxl)."
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    rows = [r for r in rows if r and any(c is not None and str(c).strip() for c in r)]

    if not rows:
        raise ValueError("La primera hoja del Excel está vacía.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return _parse_rows(filename, "xlsx", headers, rows[1:])
